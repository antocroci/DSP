#!/usr/bin/env python3
"""
Análisis DOA/TDOA Completo

"""

import argparse
import sys
import os
import numpy as np
import matplotlib.pyplot as plt
from pathlib import Path
import json
from typing import Dict, List, Optional
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# Importar módulos propios
from simulacion import SimuladorDOA, crear_senal_prueba
from tdoa import EstimadorTDOA
from doa import EstimadorDOA
from evaluacion import EvaluadorDOA


print("=== SIMULADOR DOA CON RUIDO AMBIENTE - PRUEBA ===")

print("\n1. PARÁMETROS ACÚSTICOS:")
ambiente_tipo = input("   Tipo de ambiente (1=Anecoico, 2=Reverberante) [default: 2]: ") or "2"

# Crear simulador
sim = SimuladorDOA(fs=16000)

if ambiente_tipo == "1":
    # Ambiente anecoico
    sim.simular_ambiente_anecoico(
        room_size=[float(input("   Ancho del recinto (X) en metros [default: 6.0]: ") or 6.0), 
                    float(input("   Largo del recinto (Y) en metros [default: 4.0]: ") or 4.0),
                    float(input("   Altura del recinto (Z) en metros [default: 3.0]: ") or 3.0)],
                    max_order = 0,
                    absorption = float(input("Ingrese el coeficiente de absorción de la sala (default 1): ") or 1),
                    air_absorption=False
        
    )
else:
    # Ambiente reverberante CON ruido ambiente
    sim.simular_ambiente_reverberante(
        room_size=[float(input("   Ancho del recinto (X) en metros [default: 6.0]: ") or 6.0), 
                    float(input("   Largo del recinto (Y) en metros [default: 4.0]: ") or 4.0),
                    float(input("   Altura del recinto (Z) en metros [default: 3.0]: ") or 3.0)], 
        rt60= float(input("   RT60 en segundos [default: 0.3]: ") or 0.3),
    )

# Crear señal de prueba
# signal = crear_senal_prueba("chirp", duracion=1.0)

# Cargar señal WAV individual directamente sin metadata
wav_path = input("Ingrese la ruta completa del archivo WAV para cargar (ejemplo: simulaciones/p227_004.wav): ")

from scipy.io import wavfile
import os

while not wav_path or not os.path.isfile(wav_path):
    print("Ruta inválida o archivo no encontrado. Por favor, ingrese una ruta válida.")
    wav_path = input("Ingrese la ruta completa del archivo WAV para cargar (ejemplo: simulaciones/p227_004.wav): ")

fs_loaded, signal = wavfile.read(wav_path)
signal = signal.astype(np.float32) / 32767.0  # Normalizar

# Actualizar frecuencia de muestreo si es diferente
if fs_loaded != sim.fs:
    print(f"Advertencia: La frecuencia de muestreo cargada ({fs_loaded}) es diferente de la configurada ({sim.fs}). Se actualizará.")
    sim.fs = fs_loaded

# Configuración de fuentes
print("\n2. FUENTES SONORAS:")
azimuth_real = float(input(f"     Azimuth en grados [default: 60]: ") or 60.0)
distancia_real = float(input("     Distancia en metros [default: 2.0]: ") or 2.0)
if distancia_real <= 2.0:
    print("Se debe mantener la condición de campo lejano con una distancia mayor a 2 metros.")
    distancia_real = float(input("Ingrese una distancia en metros mayor a 2.0: ") or 2.0)
    if distancia_real <= 2.0:
        distancia_real = 2.0
        print("ERROR: Valor ingresado menor a 2 metros. Distancia en metros: 2.0")
elevacion_real = float(input("     Elevación en grados [default: 0.0]: ") or 0.0)

sim.agregar_fuente(signal, azimuth=azimuth_real, distance=distancia_real, elevation=elevacion_real)

# Simular propagación
sim.simular_propagacion(agregar_ruido=True, snr_db=20)
    
# Guardar y visualizar
sim.visualizar_setup()
sim.visualizar_geometria_detallada()
    
print("\n¡Simulación completada exitosamente!")

# TDOA
print("\n=== TESTING MÓDULO TDOA ===")

# Crear instancia del estimador TDOA
estimador_tdoa = EstimadorTDOA(fs=sim.fs)

# Obtener señales de los micrófonos simulados
mic_signals = sim.signals['mic_signals']
num_mics = mic_signals.shape[0]

print(f"Analizando señales de {num_mics} micrófonos...")

# Preguntar al usuario qué método TDOA usar
print("\nMétodos TDOA disponibles:")
print("1. Correlación cruzada clásica")
print("2. GCC básico")
print("3. GCC-PHAT (recomendado)")
print("4. GCC-SCOT")
metodo_seleccionado = input("Seleccione método [default: 3]: ") or "3"

if metodo_seleccionado == "1":
    metodo_tdoa = "correlacion"
elif metodo_seleccionado == "2":
    metodo_tdoa = "gcc"
elif metodo_seleccionado == "4":
    metodo_tdoa = "gcc_scot"
else:
    metodo_tdoa = "gcc_phat"

print(f"\nUsando método TDOA: {metodo_tdoa}")

# Calcular TDOA para todos los pares de micrófonos explícitamente
pares_microfonos = [
    (0, 1),
    (0, 2),
    (0, 3),
    (1, 2),
    (1, 3),
    (2, 3)
]

resultados_tdoa = {}

print("\nCalculando TDOA para todos los pares de micrófonos:")
for i, j in pares_microfonos:
    resultado = estimador_tdoa.estimar_tdoa_par(
        mic_signals[i], 
        mic_signals[j], 
        metodo = metodo_tdoa
    )
    key = f"mic_{i+1}_mic_{j+1}"
    resultados_tdoa[key] = resultado
    tdoa_ms = resultado['tdoa_seconds'] * 1000  # Convertir a milisegundos
    print(f"  {key}: {tdoa_ms:.3f} ms (confianza: {resultado['confidence']:.3f})")

# DOA - ANÁLISIS AVANZADO
print("\n=== ANÁLISIS DOA AVANZADO ===")

# Crear instancia del estimador DOA avanzado
estimador_doa = EstimadorDOA(c=343.0)

# Obtener información del array
spacing = sim.array_geometry['spacing']
array_positions = sim.array_geometry['positions']
array_tipo = sim.array_geometry.get('tipo', 'lineal')

print(f"Configuración del array:")
print(f"  Tipo: {array_tipo}")
print(f"  Espaciado: {spacing*100:.1f} cm")
print(f"  Número de micrófonos: {num_mics}")

# Determinar geometría para el estimador DOA
if array_tipo == 'lineal':
    geometria_doa = 'linear'
elif array_tipo == 'circular':
    geometria_doa = 'circular'
else:
    geometria_doa = 'arbitrary'

# Calcular ángulos DOA usando el estimador avanzado
print(f"\nCalculando DOA con geometría: {geometria_doa}")

if geometria_doa == 'arbitrary':
    resultados_doa = estimador_doa.calcular_angulo_arribo(
        resultados_tdoa, 
        spacing, 
        geometria=geometria_doa,
        mic_positions=array_positions
    )
else:
    resultados_doa = estimador_doa.calcular_angulo_arribo(
        resultados_tdoa, 
        spacing, 
        geometria=geometria_doa
    )

print("\nResultados DOA individuales:")
angulos_estimados = []
for par, resultado in resultados_doa.items():
    if resultado.get('valido', False):
        angulo = resultado['angulo_deg']
        uncertainty = resultado.get('uncertainty_deg', 0)
        confidence = resultado.get('confidence', 1.0)
        angulo_transformado = 90 - angulo
        angulos_estimados.append(angulo_transformado)
        print(f"  {par}: {angulo_transformado:.2f}° ± {uncertainty:.2f}° (confianza: {confidence:.3f})")
    else:
        print(f"  {par}: Estimación no válida - {resultado.get('error', 'Error desconocido')}")

# Promediado de ángulos
print("\n=== PROMEDIADO DE ÁNGULOS ===")

if len(angulos_estimados) > 1:
    print("Métodos de promediado disponibles:")
    print("1. Promedio circular (recomendado)")
    print("2. Promedio aritmético")
    print("3. Promedio ponderado por confianza")
    
    metodo_promedio_sel = input("Seleccione método [default: 1]: ") or "1"
    
    if metodo_promedio_sel == "1":
        metodo_promedio = "circular"
    elif metodo_promedio_sel == "2":
        metodo_promedio = "aritmetico"
    else:
        metodo_promedio = "ponderado"
    
    # Calcular promedio
    resultado_promedio = estimador_doa.promediar_angulos(resultados_doa, metodo=metodo_promedio)
    
    if resultado_promedio.get('valido', False):
        angulo_promedio = resultado_promedio['angulo_promedio_deg']
        std_promedio = resultado_promedio['std_deg']
        num_est = resultado_promedio['num_estimaciones']
        
        print(f"\nÁngulo DOA promedio ({metodo_promedio}): {angulo_promedio:.2f}° ± {std_promedio:.2f}°")
        print(f"Basado en {num_est} estimaciones válidas")
        print(f"Ángulo real de la fuente: {azimuth_real:.2f}°")
        
        error_absoluto = abs(angulo_promedio - azimuth_real)
        print(f"Error absoluto: {error_absoluto:.2f}°")
        
        # Evaluar calidad de la estimación
        if error_absoluto < 5.0:
            print("✓ Estimación EXCELENTE (error < 5°)")
        elif error_absoluto < 10.0:
            print("✓ Estimación BUENA (error < 10°)")
        elif error_absoluto < 20.0:
            print("⚠ Estimación ACEPTABLE (error < 20°)")
        else:
            print("❌ Estimación POBRE (error > 20°)")
    else:
        print("❌ No se pudo calcular promedio válido")
        angulo_promedio = None
        resultado_promedio = None

elif len(angulos_estimados) == 1:
    angulo_promedio = angulos_estimados[0]
    print(f"\nÁngulo DOA (única estimación): {angulo_promedio:.2f}°")
    print(f"Ángulo real de la fuente: {azimuth_real:.2f}°")
    print(f"Error absoluto: {abs(angulo_promedio - azimuth_real):.2f}°")
    resultado_promedio = {'angulo_promedio_deg': angulo_promedio,
                            'valido': True,
                            'std_deg': 0.0,
                            'num_estimaciones': 1,
                        'metodo_promedio': 'unica_estimacion',}
else:
    print("❌ No se obtuvieron estimaciones válidas de DOA")
    angulo_promedio = None
    resultado_promedio = None

# Análisis de ambigüedad
print("\n=== ANÁLISIS DE AMBIGÜEDAD ===")
ambiguedad = estimador_doa.evaluar_ambiguedad(resultados_doa)

print(f"Dispersión de estimaciones: {ambiguedad['dispersion']:.2f}°")
print(f"Rango de estimaciones: {ambiguedad['rango']:.2f}°")
print(f"¿Tiene ambigüedad?: {'SÍ' if ambiguedad['tiene_ambiguedad'] else 'NO'}")

if ambiguedad['tiene_ambiguedad']:
    print("⚠ ADVERTENCIA: Se detectó ambigüedad en las estimaciones")
    print("  Esto puede deberse a:")
    print("  - Ruido excesivo")
    print("  - Reverberación")
    print("  - Configuración subóptima del array")
else:
    print("✓ Las estimaciones son consistentes")

# Triangulación de la fuente (si hay suficientes TDOAs)
print("\n=== TRIANGULACIÓN DE LA FUENTE ===")

if len(resultados_tdoa) >= 3:
    print("Intentando triangular posición de la fuente...")
    
    resultado_triangulacion = estimador_doa.triangular_fuente(
        resultados_tdoa, 
        array_positions, 
        metodo='least_squares'
    )
    
    if resultado_triangulacion.get('valido', False):
        pos_estimada = resultado_triangulacion['posicion']
        error_rms = resultado_triangulacion['error_rms']
        
        print(f"Posición estimada: [{pos_estimada[0]:.2f}, {pos_estimada[1]:.2f}, {pos_estimada[2]:.2f}] m")
        print(f"Error RMS: {error_rms:.4f} s")
        
        # Comparar con posición real de la fuente
        if sim.source_positions:
            pos_real = sim.source_positions[0]['position']
            error_posicion = np.linalg.norm(np.array(pos_estimada) - np.array(pos_real))
            print(f"Posición real: [{pos_real[0]:.2f}, {pos_real[1]:.2f}, {pos_real[2]:.2f}] m")
            print(f"Error de posición: {error_posicion:.2f} m")
    else:
        print(f"❌ Triangulación falló: {resultado_triangulacion.get('error', 'Error desconocido')}")
else:
    print("❌ Insuficientes TDOAs para triangulación (se necesitan al menos 3)")

# Visualizaciones
print("\n=== VISUALIZACIONES ===")

# Visualizar correlación TDOA
print("Mostrando correlación TDOA...")
par_visualizar = list(resultados_tdoa.keys())[0]  # Primer par
estimador_tdoa.visualizar_correlacion(
    resultados_tdoa[par_visualizar], 
    titulo=f"Correlación TDOA usando {metodo_tdoa}"
)

# Visualizar estimaciones DOA
print("Mostrando estimaciones DOA...")
estimador_doa.visualizar_estimaciones(resultados_doa, angulo_real=azimuth_real)

# Comparación de métodos TDOA (opcional)
print("\n¿Desea comparar diferentes métodos TDOA? (s/n) [default: n]: ")
comparar = input() or "n"

if comparar.lower() == "s":
    print("Comparando métodos TDOA para micrófonos 1 y 2...")
    metodos_comparar = ['correlacion', 'gcc', 'gcc_phat','gcc_scot']
    resultados_comparacion = estimador_tdoa.comparar_metodos(
        mic_signals[0], 
        mic_signals[1], 
        metodos=metodos_comparar
    )
    
    # Mostrar resultados de la comparación
    print("\nComparación de métodos TDOA:")
    for metodo, resultado in resultados_comparacion.items():
        if resultado:
            tdoa_ms = resultado['tdoa_seconds'] * 1000
            print(f"  {metodo}: {tdoa_ms:.3f} ms (confianza: {resultado['confidence']:.3f})")
    
    # Visualizar comparación
    plt.figure(figsize=(12, 8))
    for i, (metodo, resultado) in enumerate(resultados_comparacion.items()):
        if resultado:
            plt.subplot(len(resultados_comparacion), 1, i+1)
            lags_ms = resultado['lags'] * 1000 / sim.fs  # Convertir a ms
            plt.plot(lags_ms, resultado['correlation_normalized'])
            plt.axvline(x=resultado['tdoa_seconds']*1000, color='red', linestyle='--',
                        label=f'TDOA = {resultado["tdoa_seconds"]*1000:.2f} ms')
            plt.title(f'Método: {metodo}')
            plt.xlabel('Retardo (ms)')
            plt.ylabel('Correlación Normalizada')
            plt.grid(True, alpha=0.3)
            plt.legend()
    
    plt.tight_layout()
    plt.show()

# Generar reporte final
print("\n=== REPORTE FINAL ===")

if resultado_promedio:
    reporte = estimador_doa.generar_reporte(
        resultados_doa, 
        resultado_promedio, 
        angulo_real=azimuth_real
    )
    print(reporte)

# Resumen de rendimiento
print("\n=== RESUMEN DE RENDIMIENTO ===")
print(f"Configuración del experimento:")
print(f"  - Ambiente: {'Anecoico' if ambiente_tipo == '1' else 'Reverberante'}")
print(f"  - Array: {array_tipo} con {num_mics} micrófonos")
print(f"  - Espaciado: {spacing*100:.1f} cm")
print(f"  - Método TDOA: {metodo_tdoa}")
if resultado_promedio:
    print(f"  - Método promedio: {resultado_promedio.get('metodo_promedio', 'N/A')}")

print(f"\nResultados:")
print(f"  - Ángulo real: {azimuth_real:.2f}°")
if angulo_promedio is not None:
    print(f"  - Ángulo estimado: {angulo_promedio:.2f}°")
    print(f"  - Error absoluto: {abs(angulo_promedio - azimuth_real):.2f}°")
    print(f"  - Estimaciones válidas: {len(angulos_estimados)}/{len(resultados_doa)}")
else:
    print(f"  - No se pudo estimar el ángulo")

print(f"  - Ambigüedad detectada: {'SÍ' if ambiguedad['tiene_ambiguedad'] else 'NO'}")

print("\n=== Análisis DOA/TDOA completado ===")

# --- Export results and parameters to Excel ---

excel_filename = "DOA_TDOA_results.xlsx"
sheet_name = datetime.now().strftime("Run_%Y%m%d_%H%M%S")

# Collect parameters
params = {
    "ambiente_tipo": ambiente_tipo,
    "room_size_x": sim.room_size[0] if hasattr(sim, 'room_size') else None,
    "room_size_y": sim.room_size[1] if hasattr(sim, 'room_size') else None,
    "room_size_z": sim.room_size[2] if hasattr(sim, 'room_size') else None,
    "max_order": getattr(sim, 'max_order', None),
    "absorption": getattr(sim, 'absorption', None),
    "air_absorption": getattr(sim, 'air_absorption', None),
    "fs": sim.fs,
    "wav_path": wav_path,
    "azimuth_real": azimuth_real,
    "distancia_real": distancia_real,
    "elevacion_real": elevacion_real,
    "num_mics": num_mics,
    "array_tipo": array_tipo,
    "spacing_cm": spacing*100,
    "metodo_tdoa": metodo_tdoa,
    "metodo_promedio": resultado_promedio.get('metodo_promedio', None) if resultado_promedio else None,
    "angulo_real": azimuth_real,
    "angulo_estimado": angulo_promedio,
    "error_absoluto": abs(angulo_promedio - azimuth_real) if angulo_promedio is not None else None,
    "ambiguedad_detectada": ambiguedad['tiene_ambiguedad'],
}

# Convert TDOA results to DataFrame
tdoa_df = pd.DataFrame.from_dict(resultados_tdoa, orient='index')
tdoa_df = tdoa_df.rename(columns={
    'tdoa_seconds': 'TDOA_seconds',
    'confidence': 'Confidence',
    'lags': 'Lags',
    'correlation_normalized': 'Correlation_Normalized'
})

# Convert DOA results to DataFrame
doa_list = []
for par, res in resultados_doa.items():
    entry = {
        "Pair": par,
        "Valid": res.get('valido', False),
        "Angle_deg": res.get('angulo_deg', None),
        "Uncertainty_deg": res.get('uncertainty_deg', None),
        "Confidence": res.get('confidence', None),
        "Error": res.get('error', None)
    }
    doa_list.append(entry)
doa_df = pd.DataFrame(doa_list)

# Ambiguity data
ambiguity_df = pd.DataFrame([ambiguedad])

# Triangulation data
triangulation_data = {}
if 'resultado_triangulacion' in locals() and resultado_triangulacion.get('valido', False):
    triangulation_data = {
        "Pos_x": resultado_triangulacion['posicion'][0],
        "Pos_y": resultado_triangulacion['posicion'][1],
        "Pos_z": resultado_triangulacion['posicion'][2],
        "Error_RMS": resultado_triangulacion['error_rms']
    }
triangulation_df = pd.DataFrame([triangulation_data]) if triangulation_data else pd.DataFrame()

# Create a Pandas Excel writer using openpyxl engine
import zipfile

if os.path.exists(excel_filename):
    try:
        book = load_workbook(excel_filename)
        writer = pd.ExcelWriter(
            excel_filename,
            engine='openpyxl',
            mode='a',
            if_sheet_exists='new',
            engine_kwargs={'book': book}
        )
    except (zipfile.BadZipFile, IOError):
        print(f"Warning: Existing file '{excel_filename}' is corrupted or not a valid Excel file. Creating a new file.")
        writer = pd.ExcelWriter(excel_filename, engine='openpyxl')
else:
    writer = pd.ExcelWriter(excel_filename, engine='openpyxl')

# Write parameters
params_df = pd.DataFrame(list(params.items()), columns=['Parameter', 'Value'])
params_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)

# Write TDOA results
tdoa_df.to_excel(writer, sheet_name=sheet_name, index=True, startrow=len(params_df)+2)

# Write DOA results
doa_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=len(params_df)+len(tdoa_df)+4)

# Write Ambiguity results
ambiguity_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=len(params_df)+len(tdoa_df)+len(doa_df)+6)

# Write Triangulation results if available
if not triangulation_df.empty:
    triangulation_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=len(params_df)+len(tdoa_df)+len(doa_df)+len(ambiguity_df)+8)

writer.close()

print(f"\nResults and parameters exported to Excel file '{excel_filename}' in sheet '{sheet_name}'.")

# --- Save and embed matplotlib figures as images in Excel ---

from openpyxl.drawing.image import Image as XLImage
import tempfile

def save_fig_to_tempfile(fig, prefix):
    tmp_file = tempfile.NamedTemporaryFile(suffix=".png", prefix=prefix, delete=False)
    fig.savefig(tmp_file.name, bbox_inches='tight')
    tmp_file.close()
    return tmp_file.name

# List to hold image file paths for cleanup
image_files = []

# Save correlation TDOA figure
fig_corr = plt.figure(figsize=(8, 6))
par_visualizar = list(resultados_tdoa.keys())[0]  # Primer par
estimador_tdoa.visualizar_correlacion(
    resultados_tdoa[par_visualizar], 
    titulo=f"Correlación TDOA usando {metodo_tdoa}"
)
fig_corr = plt.gcf()
img_corr_path = save_fig_to_tempfile(fig_corr, "tdoa_correlation_")
image_files.append(img_corr_path)
plt.close(fig_corr)

# Save DOA estimations figure
fig_doa = plt.figure(figsize=(8, 6))
estimador_doa.visualizar_estimaciones(resultados_doa, angulo_real=azimuth_real)
fig_doa = plt.gcf()
img_doa_path = save_fig_to_tempfile(fig_doa, "doa_estimations_")
image_files.append(img_doa_path)
plt.close(fig_doa)

# Save TDOA methods comparison figure if generated
if 'comparar' in locals() and comparar.lower() == "s":
    fig_comp = plt.figure(figsize=(12, 8))
    for i, (metodo, resultado) in enumerate(resultados_comparacion.items()):
        if resultado:
            plt.subplot(len(resultados_comparacion), 1, i+1)
            lags_ms = resultado['lags'] * 1000 / sim.fs  # Convertir a ms
            plt.plot(lags_ms, resultado['correlation_normalized'])
            plt.axvline(x=resultado['tdoa_seconds']*1000, color='red', linestyle='--',
                        label=f'TDOA = {resultado["tdoa_seconds"]*1000:.2f} ms')
            plt.title(f'Método: {metodo}')
            plt.xlabel('Retardo (ms)')
            plt.ylabel('Correlación Normalizada')
            plt.grid(True, alpha=0.3)
            plt.legend()
    plt.tight_layout()
    img_comp_path = save_fig_to_tempfile(fig_comp, "tdoa_comparison_")
    image_files.append(img_comp_path)
    plt.close(fig_comp)
else:
    img_comp_path = None

# Embed images into Excel sheet
ws = writer.book[sheet_name]

# Starting row for images (after data)
start_row = len(params_df) + len(tdoa_df) + len(doa_df) + len(ambiguity_df) + (len(triangulation_df) if not triangulation_df.empty else 0) + 12
img_row = start_row

def add_image_to_sheet(ws, img_path, row, col=1):
    img = XLImage(img_path)
    img.anchor = ws.cell(row=row, column=col).coordinate
    ws.add_image(img)

# Add correlation TDOA image
add_image_to_sheet(ws, img_corr_path, img_row)
img_row += 20  # Adjust row spacing for next image

# Add DOA estimations image
add_image_to_sheet(ws, img_doa_path, img_row)
img_row += 20

# Add TDOA comparison image if available
if img_comp_path:
    add_image_to_sheet(ws, img_comp_path, img_row)
    img_row += 20

# Save and close writer again to include images
writer.close()

# Cleanup temporary image files
import os
for img_file in image_files:
    try:
        os.remove(img_file)
    except Exception as e:
        print(f"Warning: Could not remove temporary image file {img_file}: {e}")
